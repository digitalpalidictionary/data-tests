import csv
import openpyxl as xl
import re
from xlsxwriter.workbook import Workbook
import logging
import excel2json

#setup logger
logging.basicConfig(format='%(asctime)s %(message)s', datefmt='%I:%M:%S')


logging.warning(f"reading csv_file")
csv_file = "/home/bhikkhu/Bodhirasa/Dropbox/Pāli English Dictionary/Pāli English Dictionary.csv"
df = read_csv (csv_file, sep="\t")

# # convert csv to excel

# convert = input (f"convert csv to excel (y/n)? ")
# if convert == "y":
#     logging.warning(f"converting csv to excel sheet")
#     csv_file = "/home/bhikkhu/Bodhirasa/Dropbox/Pāli English Dictionary/Pāli English Dictionary-full.csv"
#     xlsx_file = "/home/bhikkhu/Bodhirasa/Dropbox/Pāli English Dictionary/Pāli English Dictionary-full.xlsx"
#     workbook = Workbook(xlsx_file)
#     worksheet = workbook.add_worksheet()
#     csv_reader = csv.reader(open(csv_file,'rt'),delimiter="\t")

#     for row, data in enumerate(csv_reader):
#         worksheet.write_row(row, 0, data)
#     workbook.close()
    
# if convert == "n":
#     pass

# open excel sheet
logging.warning ("loading excel sheet")
wb = xl.load_workbook("Pāli English Dictionary-full.xlsx")
sheet = wb["Sheet1"]
pali_last_row = sheet.max_row + 1

#convert excel to json
logging.warning ("coverting excel sheet to json")
excel2json.convert_from_file('json_test.xls')


#open txt file
logging.warning ("opening txt file")
txt_file = open ("test_results.txt", 'w', encoding= "'utf-8")



def search_for(
	test_name,
	search1_column,
	search1_string,
	search2_column,
	search2_string,
	search3_column,
	search3_string,
	search4_column,
	search4_string,
	print2_column,
	print3_column,
	print4_column,
	exceptions,
	iterations):

	logging.warning (f"test for {test_name}")

	line_break = "~" * 40

	txt_file.write (f"{line_break}\n")
	txt_file.write (f"{test_name}\n")
	txt_file.write (f"{line_break}\n")

	row_number = 2
	count = 0

	for row in range(2, pali_last_row):
		while count < int(iterations):
			search1_cell = sheet.cell(row_number, int(search1_column))
			search1_cell_value = str(search1_cell.value)
			search1 = re.search(search1_string, search1_cell_value)

			if search2_string != "":
				search2_cell = sheet.cell(row_number, int(search2_column))
				search2_cell_value = str(search2_cell.value)
				search2 = re.search(search2_string, search2_cell_value)
			else:
				search2 = 0

			if search3_string != "":
				search3_cell = sheet.cell(row_number, int(search3_column))
				search3_cell_value = str(search3_cell.value)
				search3 = re.search(search3_string, search3_cell_value)
			else:
				search3 = 0

			if search4_string != "":
				search4_cell = sheet.cell(row_number, int(search4_column))
				search4_cell_value = str(search4_cell.value)
				search4 = re.search(search4_string, search4_cell_value)
			else:
				search4 = 0
			
			headword_cell = sheet.cell(row_number, 1)
			headword_cell_value = str(headword_cell.value)

			if print2_column != "":
				print2_cell = sheet.cell(row_number, int(print2_column))
				print2_cell_value = str(print2_cell.value)
			else:
				print2_cell_value = ""

			if print3_column != "":
				print3_cell = sheet.cell(row_number, int(print3_column))
				print3_cell_value = str(print3_cell.value)
			else:
				print3_cell_value = ""

			if print4_column != "":
				print4_cell = sheet.cell(row_number, int(print4_column))
				print4_cell_value = str(print4_cell.value)
			else:
				print4_cell_value = ""

			print_seperator = " || "
			
				
			if (test_name != None
			and search1 != None 
			and search2 != None
			and search3 != None
			and search4 != None
			and headword_cell_value not in exceptions
			):
				txt_file.write(f"{headword_cell_value}{print_seperator}{print2_cell_value}{print_seperator}{print3_cell_value}{print_seperator}{print4_cell_value}\n")
				count += 1
			row_number += 1


search_for(
	test_name = "anā aṇā is not abstr",
	search1_column = "29",
	search1_string = "(anā|aṇā)",
	search2_column = "5",
	search2_string = "^((?!abstr).)*$",
	search3_column = "",
	search3_string = "",
	search4_column = "",
	search4_string = "",
	print2_column = "11",
	print3_column = "50",
	print4_column = "",
	exceptions = ["adhikuṭṭanā"],
	iterations = "3"
	)

search_for(
	test_name = "ti is not abstr",
	search1_column = "4", 
	search1_string = "fem",
	search2_column = "29",
	search2_string = "^ti$",
	search3_column = "5",
	search3_string = "(?s)^((?!abstr).)*$",
	search4_column = "",
	search4_string = "",
	print2_column = "4", 
	print3_column = "11",
	print4_column = "50",
	exceptions = [""],
	iterations = "3"
	)

search_for(
	test_name = "ā fem is not abstr",
	search1_column = "4", 
	search1_string = "fem",
	search2_column = "29",
	search2_string = "^ā$",
	search3_column = "5",
	search3_string = "(?s)^((?!abstr).)*$",
	search4_column = "",
	search4_string = "",
	print2_column = "4", 
	print3_column = "11",
	print4_column = "50",
	exceptions = ["akkharā", "ajā", "ativisā", "sutā"],
	iterations = "3"
	)

search_for(
	test_name = "tā tta tha is not abst",
	search1_column = "29", 
	search1_string = "(tā|tta|tha)$",
	search2_column = "4",
	search2_string = "(f|nt)",
	search3_column = "5",
	search3_string = "(?s)^((?!abstr).)*$",
	search4_column = "",
	search4_string = "",
	print2_column = "4", 
	print3_column = "11",
	print4_column = "50",
	exceptions = [""],
	iterations = "3"
	)

search_for(
	test_name = "ana|aṇa|anā|aṇā is not act",
	search1_column = "4", 
	search1_string = "nt",
	search2_column = "29",
	search2_string = "ana|aṇa|anā|aṇā",
	search3_column = "5",
	search3_string = "(?s)^((?!act).)*$",
	search4_column = "",
	search4_string = "",
	print2_column = "4", 
	print3_column = "11",
	print4_column = "50",
	exceptions = [""],
	iterations = "3"
	)

search_for(
	test_name = "case no constr line 2",
	search1_column = "5", 
	search1_string = " (nom|acc|instr|dat|abl|gen|loc|voc)",
	search2_column = "27",
	search2_string = "(?s)^((?!<br).)*$",
	search3_column = "",
	search3_string = "",
	search4_column = "",
	search4_string = "",
	print2_column = "4", 
	print3_column = "11",
	print4_column = "50",
	exceptions = ["aghato"],
	iterations = "5"
	)

# search_for(
# 	test_name = "",
# 	search1_column = "", 
# 	search1_string = "",
# 	search2_column = "",
# 	search2_string = "",
# 	search3_column = "",
# 	search3_string = "",
# 	search4_column = "",
# 	search4_string = "",
# 	print2_column = "4", 
# 	print3_column = "11",
# 	print4_column = "50",
# 	exceptions = [""],
# 	iterations = "5"
# 	)

# search_for(
# 	test_name = "",
# 	search1_column = "", 
# 	search1_string = "",
# 	search2_column = "",
# 	search2_string = "",
# 	search3_column = "",
# 	search3_string = "",
# 	search4_column = "",
# 	search4_string = "",
# 	print2_column = "4", 
# 	print3_column = "11",
# 	print4_column = "50",
# 	exceptions = [""],
# 	iterations = "5"
# 	)

logging.warning (f"saved to {txt_file.name}")
logging.warning ("fin")

# 1	Pāli1
# 2	Pāli2
# 3	Fin
# 4	POS
# 5	Grammar
# 6	Derived from
# 7	Neg
# 8	Verb
# 9	Trans
# 10	Case
# 11	Meaning IN CONTEXT
# 12	Literal Meaning
# 13	Non IA
# 14	Sanskrit
# 15	Sk Root
# 16	Sk Root Mn
# 17	Cl
# 18	Pāli Root
# 19	Root In Comps
# 20	V
# 21	Grp
# 22	Sgn
# 23	Root Meaning
# 24	Base
# 25	Family
# 26	Family2
# 27	Construction
# 28	Derivative
# 29	Suffix
# 30	Phonetic Changes
# 31	Compound
# 32	Compound Construction
# 33	Non-Root In Comps
# 34	Source1
# 35	Sutta1
# 36	Example1
# 37	Source 2
# 38	Sutta2
# 39	Example 2
# 40	Antonyms
# 41	Synonyms – different word
# 42	Variant – same constr or diff reading
# 43	Commentary
# 44	Notes
# 45	Cognate
# 46	Category
# 47	Link
# 48	Stem
# 49	Pattern
# 50	Buddhadatta
# 51	22
# 52	Pāli1 ≠ const
# 53	test dupl
# 54	Metadata
